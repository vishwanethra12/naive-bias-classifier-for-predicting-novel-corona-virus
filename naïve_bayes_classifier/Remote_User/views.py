from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl

import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.ensemble import VotingClassifier
from sklearn.tree import DecisionTreeClassifier
# Create your views here.
from Remote_User.models import ClientRegister_Model,Predicting_Novel_Coronavirus,detection_ratio,detection_accuracy

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('ViewYourProfile')
        except:
            pass

    return render(request,'RUser/login.html')

def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def View_primary_stage_of_diabetes_prediction(request):
    excel_file = ("Predicting_the_Novel_Coronavirus.xlsx")
    # you may put validations here to check extension or file size
    wb = openpyxl.load_workbook(excel_file)
    # getting all sheets
    sheets = wb.sheetnames
    print(sheets)
    # getting a particular sheet
    worksheet = wb["Sheet1"]
    print(worksheet)
    # getting active sheet
    active_sheet = wb.active
    print(active_sheet)
    # reading a cell
    print(worksheet["A1"].value)
    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
            print(cell.value)
        excel_data.append(row_data)

    return render(request, 'RUser/View_primary_stage_of_diabetes_prediction.html', {'excel_data': excel_data})

