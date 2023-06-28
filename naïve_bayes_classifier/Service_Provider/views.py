


from django.db.models import  Count, Avg
from django.shortcuts import render, redirect
from django.db.models import Count
from django.db.models import Q
import datetime
import xlwt
from django.http import HttpResponse
import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier
import openpyxl

# Create your views here.
from Remote_User.models import ClientRegister_Model,Predicting_Novel_Coronavirus,detection_ratio,detection_accuracy


def serviceproviderlogin(request):
    if request.method  == "POST":
        admin = request.POST.get('username')
        password = request.POST.get('password')
        if admin == "Admin" and password =="Admin":
            return redirect('View_Remote_Users')

    return render(request,'SProvider/serviceproviderlogin.html')


def viewtreandingquestions(request,chart_type):
    dd = {}
    pos,neu,neg =0,0,0
    poss=None
    topic = Predicting_Novel_Coronavirus.objects.values('ratings').annotate(dcount=Count('ratings')).order_by('-dcount')
    for t in topic:
        topics=t['ratings']
        pos_count=Predicting_Novel_Coronavirus.objects.filter(topics=topics).values('names').annotate(topiccount=Count('ratings'))
        poss=pos_count
        for pp in pos_count:
            senti= pp['names']
            if senti == 'positive':
                pos= pp['topiccount']
            elif senti == 'negative':
                neg = pp['topiccount']
            elif senti == 'nutral':
                neu = pp['topiccount']
        dd[topics]=[pos,neg,neu]
    return render(request,'SProvider/viewtreandingquestions.html',{'object':topic,'dd':dd,'chart_type':chart_type})

def View_All_Primary_Stage_of_Diabetes_Prediction(request):
    excel_file = ("Predicting_the_Novel_Coronavirus.xlsx")
    # you may put validations here to check extension or file size
    wb = openpyxl.load_workbook(excel_file)
    # getting all sheets
    sheets = wb.sheetnames
    print(sheets)
    # getting a particular sheet
    worksheet = wb["Sheet1"]
    print(worksheet)
    # getting active sheet
    active_sheet = wb.active
    print(active_sheet)
    # reading a cell
    print(worksheet["A1"].value)
    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
            print(cell.value)
        excel_data.append(row_data)
    Predicting_Novel_Coronavirus.objects.all().delete()
    for r in range(1, active_sheet.max_row):
            Predicting_Novel_Coronavirus.objects.create(
            Fever=active_sheet.cell(r, 1).value,
            Tiredness=active_sheet.cell(r, 2).value,
            Dry_Cough=active_sheet.cell(r, 3).value,
            Difficulty_in_Breathing=active_sheet.cell(r, 4).value,
            Sore_Throat=active_sheet.cell(r, 5).value,
            None_Sympton=active_sheet.cell(r, 6).value,
            Pains=active_sheet.cell(r, 7).value,
            Nasal_Congestion=active_sheet.cell(r, 8).value,
            Runny_Nose=active_sheet.cell(r, 9).value,
            Diarrhea=active_sheet.cell(r, 10).value,
            None_Experiencing=active_sheet.cell(r, 11).value,
            Age_0To9=active_sheet.cell(r, 12).value,
            Age_10To19=active_sheet.cell(r, 13).value,
            Age_20To24=active_sheet.cell(r, 14).value,
            Age_25To59=active_sheet.cell(r, 15).value,
            Age_60Above=active_sheet.cell(r, 16).value,
            Gender_Female=active_sheet.cell(r, 17).value,
            Gender_Male=active_sheet.cell(r, 18).value,
            Gender_Transgender=active_sheet.cell(r, 19).value,
            Severity_Mild=active_sheet.cell(r, 20).value,
            Severity_Moderate=active_sheet.cell(r, 21).value,
            Severity_None=active_sheet.cell(r, 22).value,
            Contact_Dont_Know=active_sheet.cell(r, 23).value,
            Contact_No=active_sheet.cell(r, 24).value,
            Contact_Yes=active_sheet.cell(r, 25).value,
            Naive_Byes=active_sheet.cell(r, 26).value,
            SVM=active_sheet.cell(r, 27).value,
            Logistic_Regression=active_sheet.cell(r, 28).value,
            RandomForestClassifier=active_sheet.cell(r, 29).value,
            Decision_Tree_Classifier=active_sheet.cell(r, 30).value,
            KNeighborsClassifier=active_sheet.cell(r, 31).value,
            )

            #obj = diabetes_prediction.objects.all()
    return render(request, 'SProvider/View_All_Primary_Stage_of_Diabetes_Prediction.html', {"excel_data": excel_data})


def Primary_Stage_of_Diabetes_Prediction_Type_Ratio(request):
    detection_ratio.objects.all().delete()
    ratio = ""
    kword = '1'
    Ikword='Coronavirus Positive'
    print(kword)
    obj = Predicting_Novel_Coronavirus.objects.all().filter(Q(Naive_Byes=kword))
    obj1 = Predicting_Novel_Coronavirus.objects.all()
    count = obj.count();
    count1 = obj1.count();
    ratio = (count / count1) * 100
    if ratio != 0:
        detection_ratio.objects.create(names=Ikword, ratio=ratio)

    ratio1 = ""
    kword1 = '0'
    Ikword1 = 'Coronavirus Negative'
    print(kword1)
    obj1 = Predicting_Novel_Coronavirus.objects.all().filter(Q(Naive_Byes=kword1))
    obj11 = Predicting_Novel_Coronavirus.objects.all()
    count1 = obj1.count();
    count11 = obj11.count();
    ratio1 = (count1 / count11) * 100
    if ratio1 != 0:
        detection_ratio.objects.create(names=Ikword1, ratio=ratio1)


    obj = detection_ratio.objects.all()
    return render(request, 'SProvider/Primary_Stage_of_Diabetes_Prediction_Type_Ratio.html', {'objs': obj})


def View_Remote_Users(request):
    obj=ClientRegister_Model.objects.all()
    return render(request,'SProvider/View_Remote_Users.html',{'objects':obj})

def ViewTrendings(request):
    topic = Predicting_Novel_Coronavirus.objects.values('topics').annotate(dcount=Count('topics')).order_by('-dcount')
    return  render(request,'SProvider/ViewTrendings.html',{'objects':topic})

def negativechart(request,chart_type):
    dd = {}
    pos, neu, neg = 0, 0, 0
    poss = None
    topic = Predicting_Novel_Coronavirus.objects.values('ratings').annotate(dcount=Count('ratings')).order_by('-dcount')
    for t in topic:
        topics = t['ratings']
        pos_count = Predicting_Novel_Coronavirus.objects.filter(topics=topics).values('names').annotate(topiccount=Count('ratings'))
        poss = pos_count
        for pp in pos_count:
            senti = pp['names']
            if senti == 'positive':
                pos = pp['topiccount']
            elif senti == 'negative':
                neg = pp['topiccount']
            elif senti == 'nutral':
                neu = pp['topiccount']
        dd[topics] = [pos, neg, neu]
    return render(request,'SProvider/negativechart.html',{'object':topic,'dd':dd,'chart_type':chart_type})

def charts(request,chart_type):
    chart1 = detection_ratio.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/charts.html", {'form':chart1, 'chart_type':chart_type})

def charts1(request,chart_type):
    chart1 = detection_accuracy.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/charts1.html", {'form':chart1, 'chart_type':chart_type})

def likeschart(request,like_chart):
    charts =detection_accuracy.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/likeschart.html", {'form':charts, 'like_chart':like_chart})

def likeschart1(request,like_chart):
    charts =detection_ratio.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/likeschart1.html", {'form':charts, 'like_chart':like_chart})

def Download_Trained_DataSets(request):

    response = HttpResponse(content_type='application/ms-excel')
    # decide file name
    response['Content-Disposition'] = 'attachment; filename="TrainedData.xls"'
    # creating workbook
    wb = xlwt.Workbook(encoding='utf-8')
    # adding sheet
    ws = wb.add_sheet("sheet1")
    # Sheet header, first row
    row_num = 0
    font_style = xlwt.XFStyle()
    # headers are bold
    font_style.font.bold = True
    # writer = csv.writer(response)
    obj = Predicting_Novel_Coronavirus.objects.all()
    data = obj  # dummy method to fetch data.
    for my_row in data:
        row_num = row_num + 1

        ws.write(row_num, 0, my_row.Fever, font_style)
        ws.write(row_num, 1, my_row.Tiredness, font_style)
        ws.write(row_num, 2, my_row.Dry_Cough, font_style)
        ws.write(row_num, 3, my_row.Difficulty_in_Breathing, font_style)
        ws.write(row_num, 4, my_row.Sore_Throat, font_style)
        ws.write(row_num, 5, my_row.None_Sympton, font_style)
        ws.write(row_num, 6, my_row.Pains, font_style)
        ws.write(row_num, 7, my_row.Nasal_Congestion, font_style)
        ws.write(row_num, 8, my_row.Runny_Nose, font_style)
        ws.write(row_num, 9, my_row.Diarrhea, font_style)
        ws.write(row_num, 10, my_row.None_Experiencing, font_style)
        ws.write(row_num, 11, my_row.Age_0To9, font_style)
        ws.write(row_num, 12, my_row.Age_10To19, font_style)
        ws.write(row_num, 13, my_row.Age_20To24, font_style)
        ws.write(row_num, 14, my_row.Age_25To59, font_style)
        ws.write(row_num, 15, my_row.Age_60Above, font_style)
        ws.write(row_num, 16, my_row.Gender_Female, font_style)
        ws.write(row_num, 17, my_row.Gender_Male, font_style)
        ws.write(row_num, 18, my_row.Gender_Transgender, font_style)
        ws.write(row_num, 19, my_row.Severity_Mild, font_style)
        ws.write(row_num, 20, my_row.Severity_Moderate, font_style)
        ws.write(row_num, 21, my_row.Severity_None, font_style)
        ws.write(row_num, 22, my_row.Contact_Dont_Know, font_style)
        ws.write(row_num, 23, my_row.Contact_No, font_style)
        ws.write(row_num, 24, my_row.Contact_Yes, font_style)
        ws.write(row_num, 25, my_row.Naive_Byes, font_style)
        ws.write(row_num, 26, my_row.SVM, font_style)
        ws.write(row_num, 27, my_row.Logistic_Regression, font_style)
        ws.write(row_num, 28, my_row.RandomForestClassifier, font_style)
        ws.write(row_num, 29, my_row.Decision_Tree_Classifier, font_style)
        ws.write(row_num, 30, my_row.KNeighborsClassifier, font_style)

    wb.save(response)
    return response

def Train_Test_DataSets(request):

    detection_accuracy.objects.all().delete()

    pd.pandas.set_option('display.max_columns', None)
    pd.pandas.set_option('display.max_rows', None)

    train = pd.read_csv('Covid19.csv')
    train.head()
    data = train.copy()
    #data = data.drop(['Severity_None', 'None_Sympton', 'None_Experiencing', 'Contact_Dont_Know', 'Country', 'Contact_No'], axis=1)
    data = data.drop(['Country'], axis=1)
    data.head()
    data1 = data.copy()
    #data1 = data.drop(['Severity_Moderate', 'Severity_Mild'], axis=1)
    y_data = data1['Severity_Severe']
    x_data = data1.drop(['Severity_Severe'], axis=1)

    SEED = 42

    from sklearn.model_selection import train_test_split
    X_train, X_test, Y_train, Y_test = train_test_split(x_data, y_data, test_size=0.3, random_state=SEED)


    from sklearn.model_selection import KFold
    from sklearn.model_selection import cross_val_score
    k_fold = KFold(n_splits=10, shuffle=True, random_state=0)
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.model_selection import GridSearchCV
    from sklearn.metrics import accuracy_score
    from sklearn.metrics import classification_report
    from sklearn.metrics import accuracy_score, confusion_matrix, classification_report

    print("Naive Bayes")

    from sklearn.naive_bayes import MultinomialNB
    NB = MultinomialNB()
    NB.fit(X_train, Y_train)
    predict_nb = NB.predict(X_test)
    naivebayes = accuracy_score(Y_test, predict_nb) * 100
    print("ACCURACY")
    print(naivebayes)
    print("CLASSIFICATION REPORT")
    print(classification_report(Y_test, predict_nb))
    print("CONFUSION MATRIX")
    print(confusion_matrix(Y_test, predict_nb))
    detection_accuracy.objects.create(names="Naive Bayes", ratio=naivebayes)

    rf = RandomForestClassifier()
    rf.fit(X_train, Y_train)
    pred_rfc = rf.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(Y_test, pred_rfc) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(Y_test, pred_rfc))
    print("CONFUSION MATRIX")
    print(confusion_matrix(Y_test, pred_rfc))
    detection_accuracy.objects.create(names="Random Forest Classifier", ratio=accuracy_score(Y_test, pred_rfc) * 100)

    # SVM Model
    print("SVM")
    from sklearn import svm
    lin_clf = svm.LinearSVC()
    lin_clf.fit(X_train, Y_train)
    predict_svm = lin_clf.predict(X_test)
    svm_acc = accuracy_score(Y_test, predict_svm) * 100
    print("ACCURACY")
    print(svm_acc)
    print("CLASSIFICATION REPORT")
    print(classification_report(Y_test, predict_svm))
    print("CONFUSION MATRIX")
    print(confusion_matrix(Y_test, predict_svm))
    detection_accuracy.objects.create(names="SVM", ratio=svm_acc)

    print("KNeighborsClassifier")
    from sklearn.neighbors import KNeighborsClassifier
    kn = KNeighborsClassifier()
    kn.fit(X_train, Y_train)
    knpredict = kn.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(Y_test, knpredict) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(Y_test, knpredict))
    print("CONFUSION MATRIX")
    print(confusion_matrix(Y_test, knpredict))
    detection_accuracy.objects.create(names="KNeighborsClassifier", ratio=accuracy_score(Y_test, knpredict) * 100)

    print("Decision Tree Classifier")
    dtc = DecisionTreeClassifier()
    dtc.fit(X_train, Y_train)
    dtcpredict = dtc.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(Y_test, dtcpredict) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(Y_test, dtcpredict))
    print("CONFUSION MATRIX")
    print(confusion_matrix(Y_test, dtcpredict))
    detection_accuracy.objects.create(names="Decision Tree Classifier", ratio=accuracy_score(Y_test, dtcpredict) * 100)

    # Logistic Regression Model
    print("Logistic Regression")
    from sklearn.linear_model import LogisticRegression
    logreg = LogisticRegression(random_state=42)
    logreg.fit(X_train, Y_train)
    predict_log = logreg.predict(X_test)
    logistic = accuracy_score(Y_test, predict_log) * 100
    print("ACCURACY")
    print(logistic)
    from sklearn.metrics import confusion_matrix, f1_score
    print("CLASSIFICATION REPORT")
    print(classification_report(Y_test, predict_log))
    print("CONFUSION MATRIX")
    print(confusion_matrix(Y_test, predict_log))
    detection_accuracy.objects.create(names="Logistic Regression", ratio=logistic)


    submission_df = pd.DataFrame(columns=[])

    submission_df['Age'] = X_test.Fever
    submission_df['Tiredness'] = X_test.Tiredness
    submission_df['Dry_Cough'] = X_test.Dry_Cough
    submission_df['Difficulty_in_Breathing'] = X_test.Difficulty_in_Breathing

    submission_df['Sore_Throat'] = X_test.Sore_Throat
    submission_df['None_Sympton'] = X_test.None_Sympton

    submission_df['Pains'] = X_test.Pains
    submission_df['Nasal_Congestion'] = X_test.Nasal_Congestion

    submission_df['Runny_Nose'] = X_test.Runny_Nose
    submission_df['Diarrhea'] = X_test.Diarrhea

    submission_df['None_Experiencing'] = X_test.None_Experiencing
    submission_df['Age_0To9'] = X_test.Age_0To9

    submission_df['Age_10To19'] = X_test.Age_10To19
    submission_df['Age_20To24'] = X_test.Age_20To24
    submission_df['Age_25To59'] = X_test.Age_25To59
    submission_df['Age_60Above'] = X_test.Age_60Above



    submission_df['Gender_Female'] = X_test.Gender_Female
    submission_df['Gender_Male'] = X_test.Gender_Male

    submission_df['Gender_Transgender'] = X_test.Gender_Transgender
    submission_df['Severity_Mild'] = X_test.Severity_Mild

    submission_df['Severity_Moderate'] = X_test.Severity_Moderate
    submission_df['Severity_None'] = X_test.Severity_None

    #submission_df['Severity_Severe'] = X_test.Severity_Severe
    submission_df['Contact_Dont_Know'] = X_test.Contact_Dont_Know

    submission_df['Contact_No'] = X_test.Contact_No
    submission_df['Contact_Yes'] = X_test.Contact_Yes
    #submission_df['Country'] = X_test.Country



    submission_df['Naive Byes'] =  NB.predict(X_test)
    submission_df['SVM'] = lin_clf.predict(X_test)
    submission_df['Logistic Regression'] = logreg.predict(X_test)
    submission_df['RandomForestClassifier'] = rf.predict(X_test)
    submission_df['Decision Tree Classifier'] =dtc.predict(X_test)
    submission_df['KNeighborsClassifier'] = kn.predict(X_test)

    submission_df.to_excel('Predicting_the_Novel_Coronavirus.xlsx', index=False)

    excel_file = ("Predicting_the_Novel_Coronavirus.xlsx")
    # you may put validations here to check extension or file size
    wb = openpyxl.load_workbook(excel_file)
    # getting all sheets
    sheets = wb.sheetnames
    print(sheets)
    # getting a particular sheet
    worksheet = wb["Sheet1"]
    print(worksheet)
    # getting active sheet
    active_sheet = wb.active
    print(active_sheet)
    # reading a cell
    print(worksheet["A1"].value)
    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
            print(cell.value)
        excel_data.append(row_data)

    obj = detection_accuracy.objects.all()

    return render(request,'SProvider/Train_Test_DataSets.html', {'objs': obj,"excel_data": excel_data})














